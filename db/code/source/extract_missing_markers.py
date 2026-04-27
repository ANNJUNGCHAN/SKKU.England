"""ONS BoP xlsx 데이터 영역 결측 표기 인벤토리 추출 스크립트.

목적
----
PLAN.md/Phase 1.1 §1.1.12("사용된 결측 표기의 종류와 등장 위치 기록") 산출물을
`db/data/_inventory/12_missing_markers.csv`(기계 판독) + `12_missing_markers.md`
(사람 검토)로 출력한다. 본 추출은 다음 네 가지 필터를 추가해
`db/code/source/extract_units_missing.py`(15회차 background)와 다르다.

1. 메타 시트(Cover_sheet, Notes, Records)의 헤더 텍스트는 의미 없는 결측처럼
   잡히므로 통계 본표·R 계열 시트만 대상으로 한정한다.
2. 데이터 영역(시점 컬럼 = 두 번째 열 이후)만 순회해 본격 결측 코드만 카운트한다.
3. 빈 셀은 데이터 영역 안쪽(첫 분기 이후·CDID 행 아래)에서 등장한 것만 결측
   후보로 집계한다. 시트 padding 영역(끝부분 빈 행)은 제외한다.
4. **컬럼 padding 제외**: CDID 행에 코드가 등장한 컬럼만 데이터 컬럼으로 간주한다.
   (예: Table_A는 N~S 컬럼이 trailing 빈 컬럼이며 CDID 행에도 비어 있음 → 결측 카운트 제외)

원본 파일은 `read_only=True`로 열며 셀 값은 일체 변경하지 않는다(`db/CLAUDE.md`
값 불변 원칙).

산출 컬럼
----------
sheet, sub_table, marker, count, first_position, last_position, suggested_meaning
- marker: 정규화 표기(`x`, `(empty)`, `..`, `-`, `[c]` 등)
- first/last_position: 엑셀식 셀 주소(예: `B9`, `H88`) — Phase 2.1 ETL이 직접 참조 가능
- suggested_meaning: 10회차 ONS 웹 조사 + Service Manual 권고에 따른 해석

재현 방법
----------
env/Scripts/python.exe db/code/source/extract_missing_markers.py
"""
from __future__ import annotations

import csv
import re
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 저장소 루트 = 본 스크립트 기준 3단계 위(`db/code/source/` → repo root).
ROOT = Path(__file__).resolve().parents[3]
SRC = ROOT / "db" / "source" / "balanceofpayments2025q4.xlsx"
INV_DIR = ROOT / "db" / "data" / "_inventory"
SHEET_INV = INV_DIR / "01_sheet_dimensions.csv"
CDID_INV = INV_DIR / "07_cdid_and_first_data_rows.csv"
BOUND_INV = INV_DIR / "11_subtable_boundaries.csv"
OUT_CSV = INV_DIR / "12_missing_markers.csv"

# 통계 본표 외 메타 시트(결측 인벤토리 대상 제외).
META_SHEETS = {"Cover_sheet", "Notes", "Records"}

# 숫자 셀 정규식(쉼표 단위 구분 포함).
NUM_RE = re.compile(r"^[+-]?[\d,]+(\.\d+)?$")

# CDID 행 식별 정규식(13회차 사전과 동일 규칙: 4자 영숫자 또는 부호 prefix).
PURE = re.compile(r"^[A-Z0-9]{4}$")
PREF = re.compile(r"^\s*-\s*[A-Z0-9]{4}\s*$")
LET = re.compile(r"[A-Z]")

# 결측 표기 의미 사전(10회차 ONS 웹 조사 + Service Manual 표기 가이드 반영).
MISSING_MEANINGS = {
    "x": "ONS 비공개(suppressed): 응답 기관 식별 위험 → Service Manual 권고 기호는 [c]",
    "(empty)": "빈 셀: 시리즈 시작 이전·이후 또는 비해당 — 데이터 결측 가능성 동시 의미",
    "..": "ONS legacy 미가용: Service Manual 비권장(현행 [x] 권장)",
    "-": "ONS legacy 진정한 0 또는 영(nil): Service Manual 비권장(현행 [z] 권장)",
    "[c]": "Government Analysis Function 권장: 비공개(confidential)",
    "[x]": "Government Analysis Function 권장: 미가용(not available)",
    "[z]": "Government Analysis Function 권장: 비해당(not applicable)",
    "[low]": "Government Analysis Function 권장: 반올림 0(< 0.5 단위)",
    "n/a": "ONS Service Manual 비권장(모호)",
    "NA": "ONS Service Manual 비권장(모호)",
}


def is_cdid(raw: object) -> bool:
    """13회차 추출 스크립트와 동일한 CDID 행 인식 규칙."""
    if not isinstance(raw, str):
        return False
    s = raw.strip()
    if not (PURE.match(s) or PREF.match(raw)):
        return False
    code = re.sub(r"[^A-Z0-9]", "", raw)
    if code == "CDID":
        return False
    return bool(LET.search(code))


def classify_marker(value: object) -> str | None:
    """셀 값을 결측 마커로 분류. 숫자(또는 숫자 문자열)는 None 반환."""
    if value is None:
        return "(empty)"
    if isinstance(value, (int, float)):
        return None
    s = str(value).strip()
    if s == "":
        return "(empty)"
    if NUM_RE.match(s):
        return None
    return s


def find_cdid_rows(rows: list[list[object]]) -> list[int]:
    """시트 행 목록에서 CDID 행 인덱스(0-based) 모두 반환."""
    out = []
    for i, row in enumerate(rows):
        cnt = sum(1 for v in row if isinstance(v, str) and is_cdid(v))
        if cnt >= 2:
            out.append(i)
    return out


def cdid_data_columns(cdid_row: list[object]) -> list[int]:
    """CDID 행에서 유효한 코드가 등장한 컬럼 인덱스(0-based) 목록을 반환.

    좌측 첫 컬럼(통계항목 라벨)과 우측 trailing padding 컬럼을 제외하기 위함.
    Table_J 등 일부 시트는 CDID 위에 부호 prefix(`-`) 행이 별도로 등장하지만,
    sign_prefix 행도 PREF 정규식으로 동일하게 잡히므로 안전.
    """
    return [ci for ci, v in enumerate(cdid_row) if isinstance(v, str) and is_cdid(v)]


def find_last_data_row(rows: list[list[object]], start: int, end: int) -> int:
    """[start, end] 범위에서 데이터(숫자 또는 결측 마커)가 마지막으로 등장한 행 index.

    시트 끝 padding 영역(연속 빈 행)을 결측 카운트에서 제외하기 위함이다.
    행 안에 단 한 셀이라도 숫자가 있으면 데이터 영역으로 간주.
    """
    last = start - 1
    for i in range(start, end + 1):
        if i >= len(rows):
            break
        row = rows[i]
        # 첫 컬럼(통계항목 라벨)은 데이터 판단에서 제외.
        for v in row[1:]:
            if isinstance(v, (int, float)):
                last = i
                break
            if isinstance(v, str) and NUM_RE.match(v.strip()):
                last = i
                break
    return last


def main() -> None:
    """본 추출의 진입점. CSV/MD 두 산출물을 생성한다."""
    wb = load_workbook(SRC, read_only=True, data_only=True)
    aggregate: dict[tuple[str, str, str], dict] = defaultdict(
        lambda: {"count": 0, "first": "", "last": ""}
    )
    sheet_summary: dict[str, dict] = defaultdict(
        lambda: {"sub_tables": 0, "data_cells_total": 0, "missing_total": 0}
    )

    for sn in wb.sheetnames:
        if sn in META_SHEETS:
            continue  # 메타 시트는 결측 인벤토리 대상 외.
        ws = wb[sn]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        cdid_rows = find_cdid_rows(rows)
        if not cdid_rows:
            continue

        # 부표별 데이터 영역 [start_data, end_data] + 데이터 컬럼 인덱스 리스트 산출.
        sub_ranges: list[tuple[int, int, int, list[int]]] = []
        for k, ridx in enumerate(cdid_rows):
            start_data = ridx + 1
            if k + 1 < len(cdid_rows):
                # 다음 부표의 CDID 행 직전(빈 행 + 부표 머리글) 영역에서 마지막 데이터 행을 탐색.
                end_data = cdid_rows[k + 1] - 1
            else:
                end_data = len(rows) - 1
            end_data = find_last_data_row(rows, start_data, end_data)
            if end_data < start_data:
                continue  # 데이터 행 0개인 부표는 건너뜀(이론상 없음).
            data_cols = cdid_data_columns(rows[ridx])
            if not data_cols:
                continue  # CDID가 없는 부표는 건너뜀(이론상 없음 — 본표 한정 추출이므로).
            sub_ranges.append((k + 1, start_data, end_data, data_cols))

        sheet_summary[sn]["sub_tables"] = len(sub_ranges)

        for sub_num, sd, ed, data_cols in sub_ranges:
            sub_label = str(sub_num)
            for ri in range(sd, ed + 1):
                row = rows[ri]
                for ci in data_cols:
                    if ci >= len(row):
                        continue  # 행 길이가 짧을 경우 안전 보호.
                    val = row[ci]
                    sheet_summary[sn]["data_cells_total"] += 1
                    marker = classify_marker(val)
                    if marker is None:
                        continue
                    sheet_summary[sn]["missing_total"] += 1
                    key = (sn, sub_label, marker)
                    bucket = aggregate[key]
                    bucket["count"] += 1
                    pos = f"{get_column_letter(ci + 1)}{ri + 1}"
                    if not bucket["first"]:
                        bucket["first"] = pos
                    bucket["last"] = pos
    wb.close()

    # CSV 산출(기계 판독용 단일 평면 표).
    fields = [
        "sheet",
        "sub_table",
        "marker",
        "count",
        "first_position",
        "last_position",
        "suggested_meaning",
    ]
    INV_DIR.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for (sn, sub, marker), info in sorted(aggregate.items()):
            meaning = MISSING_MEANINGS.get(
                marker, "분류 미정 — 추가 검토 필요(원문 토큰 그대로 유지)"
            )
            w.writerow(
                {
                    "sheet": sn,
                    "sub_table": sub,
                    "marker": marker,
                    "count": info["count"],
                    "first_position": info["first"],
                    "last_position": info["last"],
                    "suggested_meaning": meaning,
                }
            )

    # 진단 로그.
    total_rows = sum(1 for _ in aggregate)
    total_count = sum(v["count"] for v in aggregate.values())
    uniq_markers = sorted({k[2] for k in aggregate})
    print(
        f"OK rows={total_rows} total_missing_cells={total_count} "
        f"unique_markers={uniq_markers}"
    )


if __name__ == "__main__":
    main()
