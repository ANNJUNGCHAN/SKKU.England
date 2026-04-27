"""
extract_latest.py — 영국 BoP 원본 xlsx 분기 시계열 말단 행 추출 스크립트
========================================================================

[목적]
    `db/source/balanceofpayments2025q4.xlsx`에서 분기 시계열을 담고 있는
    핵심 시트(Table_A, Table_B, Table_J, Table_K)의 **마지막 30행**을 그대로
    출력해, 가장 최근 분기(예: 2025 Q4)를 빠르게 확인할 수 있게 합니다.
    REPORT.md의 헤드라인·시계열 표 작성 시 사용한 1차 데이터 추출 도구입니다.

[전제]
    - 원본은 `db/source/<항목><yyyy><기간>.xlsx` 위치에 있습니다.
    - 원본을 **읽기 전용**으로만 접근하며, 값 변경 금지.
    - 가상환경: 저장소 루트 `env/`.
      필수 패키지: `openpyxl` (requirements.txt 동기화 완료).

[사용법]
    1) 가상환경 활성화 (PowerShell)
         > .\\env\\Scripts\\Activate.ps1
       또는 Git Bash
         $ source env/Scripts/activate

    2) 저장소 루트에서 실행
         (env) > python db/code/source/extract_latest.py

    3) 파일로 저장
         (env) > python db/code/source/extract_latest.py > .tail.txt

[출력 형식]
    ## <시트명>  rows=<총행수> cols=<총열수>
      r <행번호>: ['<col1>', '<col2>', ...]   ← 각 셀 최대 25자, 14열까지

[추출 대상 시트]
    - Table_A : BoP 전체 잔액 요약 (경상·자본·금융계정 부표 수직 결합)
    - Table_B : 경상수지 (£million 부표 + % of GDP 부표 포함)
    - Table_J : 금융계정 (직접·증권·파생·기타·준비자산 등)
    - Table_K : 국제투자대조표(IIP), 분기말 stock

[다른 시트를 추출하려면]
    main 영역의 SHEETS 리스트에 시트명을 추가하면 됩니다.
    예) SHEETS = ["Table_A", "Table_E", "Table_F", "Table_K"]

[주의사항]
    - Table_A 등은 1997 연간 → 분기 부표가 수직으로 결합돼 있어, 마지막 30행은
      "최근 약 8분기 + 직전 부표 꼬리"가 섞일 수 있습니다. 분기/연간 구분은
      A열(Time Period)의 표기(`'2025 Q4'` vs `'2024'`)로 판별하세요.
    - 빈 셀은 ''(공문자열)로 출력합니다.
    - 비ASCII(£ 등) 문자 보존을 위해 stdout을 UTF-8로 강제 래핑합니다.
    - 본 스크립트는 **출력 전용**으로 어떤 파일도 생성·수정하지 않습니다.

[관련 파일]
    - `db/code/source/inspect_bop.py` : 전체 시트 헤더 구조 점검
    - `db/REPORT.md`                  : 본 스크립트 출력에 기반한 분석
    - `db/CLAUDE.md`                  : `db/source/` 가공 규칙
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl

# Windows 한국어 로캘에서도 £ 등 비ASCII 문자가 깨지지 않도록 stdout을 UTF-8로 강제.
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# 저장소 루트 = 본 파일 기준 3단계 상위 (db/code/source/extract_latest.py → repo root).
REPO_ROOT = Path(__file__).resolve().parents[3]
SRC = REPO_ROOT / "db" / "source" / "balanceofpayments2025q4.xlsx"

# 추출 대상 시트와 출력할 꼬리 행 수.
SHEETS: list[str] = ["Table_A", "Table_B", "Table_J", "Table_K"]
TAIL_ROWS: int = 30
MAX_COLS_PREVIEW: int = 14
MAX_CELL_CHARS: int = 25


def dump_tail(sheet_name: str, max_tail: int = TAIL_ROWS) -> None:
    """주어진 시트의 마지막 `max_tail` 행을 그대로 출력합니다(원본 무수정)."""
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb[sheet_name]
    print(f"\n## {sheet_name}  rows={ws.max_row} cols={ws.max_column}")

    rows = list(ws.iter_rows(values_only=True))
    start_row_idx = len(rows) - max_tail + 1

    for i, row in enumerate(rows[-max_tail:], start=start_row_idx):
        # 각 셀을 최대 25자 미리보기, 빈 셀은 ''. 14열까지만 좌측 표기.
        trimmed = [
            (str(c)[:MAX_CELL_CHARS] if c is not None else "")
            for c in row[:MAX_COLS_PREVIEW]
        ]
        # 모든 셀이 공백이면 출력 생략 (시트 끝의 빈 패딩 제거).
        if any(t.strip() for t in trimmed):
            print(f"  r{i:>4}: {trimmed}")


def main() -> None:
    for sheet in SHEETS:
        dump_tail(sheet, TAIL_ROWS)


if __name__ == "__main__":
    main()
