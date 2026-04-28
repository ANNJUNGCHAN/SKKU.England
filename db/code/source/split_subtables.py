"""Phase 2.1 시트 단위 가로형 CSV 분리 ETL.

목적
----
db/CHECKLIST.md §2.1 12개 항목을 일괄 충족한다.

- 원본 시트 1개 = 시트 본체. 본표 17 시트는 부표마다 별도 CSV 1개로 분리하며,
  메타 시트 3개(Cover_sheet·Notes·Records)는 한국어 메모로 별도 분기 처리한다.
- 한 CSV에 둘 이상의 시트나 부표를 결합하지 않는다.
- 좌측 식별자 열은 시점(time_period, ECOS의 TIME 대응)이며, 첫 번째 컬럼으로 고정한다.
- 그 외 컬럼명은 ECOS의 ITEM_CODE1 대응인 ONS CDID 코드(영문 4자, sign_prefix 보존).
- 컬럼명은 ASCII 소문자·숫자·언더스코어로만 구성(시점 + cdid 정규화).
- 결측 표기(`x`, 빈 셀 등)는 원문 그대로 문자열로 보존(빈 셀은 빈 문자열로 기록).
- 셀 값 변경·반올림·치환·환산·합산·보간 일절 없음(`db/CLAUDE.md` "값 불변" 원칙).

부표 차원
----------
18회차 사전 점검(`db/data/_inventory/18_subtable_curriculum_alignment.md`)에서
정규화한 5축 부표 차원을 ETL이 읽도록 설계. 본 스크립트는 마스터 인벤토리
(`db/data/_inventory/15_master_inventory.csv`) 1행을 시트 처리 단위로 사용한다.

산출 경로 (`db/data/`)
----------------------
- 본표 부표 CSV: balanceofpayments2025q4_<table_code_lower>_sub<n>.csv (63개)
- ETL 작업 로그(요약): db/data/_etl_log/phase2_1_split_log.csv

원본 보존
----------
db/source/balanceofpayments2025q4.xlsx는 read_only=True로 열며 일절 수정하지 않는다.

재현
----
env/Scripts/python.exe db/code/source/split_subtables.py
"""
from __future__ import annotations

import csv
import re
from pathlib import Path
from openpyxl import load_workbook

# 저장소 루트 = 본 스크립트 기준 3단계 위.
ROOT = Path(__file__).resolve().parents[3]
SRC = ROOT / "db" / "source" / "balanceofpayments2025q4.xlsx"
INV = ROOT / "db" / "data" / "_inventory" / "15_master_inventory.csv"
OUT_DIR = ROOT / "db" / "data"
LOG_DIR = ROOT / "db" / "data" / "_etl_log"
LOG_CSV = LOG_DIR / "phase2_1_split_log.csv"

# CDID 행 인식 정규식(13회차 사전·12회차 결측 추출과 동일 규칙).
PURE = re.compile(r"^[A-Z0-9]{4}$")
PREF = re.compile(r"^\s*-\s*[A-Z0-9]{4}\s*$")
LET = re.compile(r"[A-Z]")

# 시점 라벨 정규식(연 YYYY 또는 분기 YYYY Qn). 다른 라벨 등장 시 원문 그대로 문자열 보존.
TIME_RE = re.compile(r"^(?P<year>\d{4})(?:\s+Q(?P<q>[1-4]))?$")


def is_cdid(raw: object) -> bool:
    """13회차 추출 스크립트와 동일한 CDID 행 인식 규칙."""
    if not isinstance(raw, str):
        return False
    s = raw.strip()
    if not (PURE.match(s) or PREF.match(raw)):
        return False
    code = re.sub(r"[^A-Z0-9]", "", raw)
    if code == "CDID":
        return False
    return bool(LET.search(code))


def normalize_cdid_column(raw: str) -> str:
    """CDID 셀 원문을 컬럼명용 ASCII 소문자 영숫자·언더스코어로 정규화.

    - 부호 prefix(` -CDID`) 행은 `_neg_<cdid>`로 변환해 부호 반전 의미를 컬럼명에 보존.
      이는 인벤토리 06회차 부호 규약 검증과 13회차 사전의 sign_prefix 컬럼 의미와 정합.
    - 일반 CDID(`HBOG` 등)는 소문자(`hbog`)로 변환.
    """
    s = raw.strip()
    is_neg = bool(PREF.match(raw))
    code = re.sub(r"[^A-Z0-9]", "", raw)
    if is_neg:
        return f"_neg_{code.lower()}"
    return code.lower()


def normalize_time_period(raw: object) -> str:
    """시점 셀(첫 컬럼)을 ECOS TIME 규약으로 정규화.

    - 정수 1997, 2025: "1997", "2025" (연간 YYYY)
    - 문자열 "1997 Q1", "2025 Q4": "1997Q1", "2025Q4" (분기 YYYYQn, ECOS 규약)
    - 그 외 라벨(`Time Period`, 빈 셀 등): 원문 그대로 문자열 보존(가공·치환 금지).
    """
    if raw is None:
        return ""
    if isinstance(raw, int):
        return str(raw)
    if isinstance(raw, float):
        # 엑셀이 정수년을 float으로 읽는 경우 정수 표현 복원(소수점 자리 0인 경우만).
        if raw.is_integer():
            return str(int(raw))
        return str(raw)
    s = str(raw).strip()
    m = TIME_RE.match(s)
    if m and m.group("q"):
        return f"{m.group('year')}Q{m.group('q')}"
    if m and not m.group("q"):
        return m.group("year")
    return s  # 비정형 라벨은 원문 보존.


def cell_to_string(value: object) -> str:
    """데이터 셀 값을 원문 보존 형태의 문자열로 변환.

    - None / 빈 문자열: "" (빈 셀)
    - int / float: 숫자 그대로 문자열화(원본 데이터 보존을 위해 자릿수 변경 없음)
    - str: 양 끝 공백 trim 후 원문 그대로

    어떤 경우에도 0/NaN으로의 치환·반올림·환산이 없다.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        # 엑셀에 boolean이 들어오는 경우 원문 표현 유지.
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        # 부동소수점 손실 방지: openpyxl이 반환한 그대로 직렬화.
        return repr(value) if "e" in repr(value) else str(value)
    return str(value).strip()


def parse_master_row(row: dict[str, str]) -> dict:
    """마스터 인벤토리 행을 ETL이 사용하는 형태로 정규화."""
    cdid_str = row["all_cdid_rows"]
    blank_str = row["blank_row_positions"]
    return {
        "sheet": row["sheet"],
        "table_code": row["table_code"],
        "classification_code": row["classification_code"],
        "n_subtables": int(row["n_subtables"]) if row["n_subtables"] else 0,
        "n_rows": int(row["n_rows"]) if row["n_rows"] else 0,
        "all_cdid_rows": [int(x) for x in cdid_str.split(";") if x] if cdid_str else [],
        "blank_rows": [int(x) for x in blank_str.split(";") if x] if blank_str else [],
        "unit_normalized": row["unit_normalized"],
        "subtable_split_required": row["subtable_split_required"],
    }


def process_subtable(
    rows: list[list[object]],
    cdid_row_1based: int,
    data_end_1based: int,
) -> tuple[list[str], list[list[str]], int]:
    """단일 부표를 가로형 CSV 행 목록으로 변환.

    Parameters
    ----------
    rows
        시트 전체 행 목록(0-based 인덱싱). openpyxl iter_rows 결과.
    cdid_row_1based
        부표 헤더 CDID 행의 1-based 행 번호.
    data_end_1based
        부표의 마지막 데이터 행의 1-based 행 번호(다음 부표 직전 빈 행을 **포함하지 않는** 마지막 데이터 행).

    Returns
    -------
    header : list[str]
        ["time_period", <cdid 컬럼 정규화 이름들...>]
    data_rows : list[list[str]]
        시점 1행 + CDID 컬럼 N개의 데이터 행 목록.
    raw_cell_count : int
        원본 셀 수(검증용). CDID 행 컬럼 헤더 셀 + 데이터 영역 셀(시점 컬럼 포함).
    """
    cdid_idx = cdid_row_1based - 1  # 0-based로 변환.
    cdid_row = rows[cdid_idx]
    # CDID 컬럼 인덱스만 추출(첫 컬럼 = 시점, 둘째 컬럼부터 CDID).
    cdid_col_indices = [
        ci for ci, v in enumerate(cdid_row) if isinstance(v, str) and is_cdid(v)
    ]
    header = ["time_period"] + [
        normalize_cdid_column(str(cdid_row[ci])) for ci in cdid_col_indices
    ]

    # 데이터 행 범위(0-based, inclusive): [cdid_idx + 1, data_end_1based - 1].
    data_start_idx = cdid_idx + 1
    data_end_idx = data_end_1based - 1
    data_rows: list[list[str]] = []
    raw_cell_count = 0
    # CDID 행 자체의 셀 개수도 검증 합계에 포함(컬럼 헤더로 사용된 원본 셀).
    raw_cell_count += 1 + len(cdid_col_indices)  # 시점 라벨 + CDID 셀들.
    for ri in range(data_start_idx, data_end_idx + 1):
        if ri >= len(rows):
            break
        row = rows[ri]
        # 시점 셀(첫 컬럼) + CDID 컬럼 N개 = 1 + N 셀이 한 데이터 행을 이룬다.
        time_str = normalize_time_period(row[0] if row else None)
        out_row = [time_str]
        for ci in cdid_col_indices:
            val = row[ci] if ci < len(row) else None
            out_row.append(cell_to_string(val))
        data_rows.append(out_row)
        raw_cell_count += 1 + len(cdid_col_indices)
    return header, data_rows, raw_cell_count


def main() -> None:
    """본 ETL 진입점. 마스터 인벤토리를 순회해 부표마다 CSV 1개씩 생성."""
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)

    # 마스터 인벤토리 로드.
    with INV.open(encoding="utf-8") as f:
        masters = [parse_master_row(r) for r in csv.DictReader(f)]

    wb = load_workbook(SRC, read_only=True, data_only=True)
    log_rows: list[dict] = []
    out_files_total = 0
    cells_total = 0

    for m in masters:
        sn = m["sheet"]
        # 메타 시트는 한국어 메모로 별도 처리(본 ETL 대상 외).
        if m["classification_code"] == "meta_notes":
            log_rows.append(
                {
                    "sheet": sn,
                    "sub_table": "",
                    "out_file": "",
                    "n_data_rows": 0,
                    "n_data_cols": 0,
                    "raw_cell_count": 0,
                    "skipped_reason": "meta_notes (한국어 메모로 별도 처리)",
                }
            )
            continue

        ws = wb[sn]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]

        cdid_rows_1based = m["all_cdid_rows"]
        blanks_1based = m["blank_rows"]
        n_total_rows_1based = m["n_rows"]
        table_code_lower = m["table_code"].replace(".", "_").lower()

        for k in range(m["n_subtables"]):
            cdid_row_1b = cdid_rows_1based[k]
            # 데이터 종료 행(1-based, inclusive). 다음 부표 직전 빈 행 또는 시트 마지막 행.
            # 빈 행은 데이터가 아니므로 종료 행은 빈 행 직전(blank - 1)이 된다.
            if k < len(blanks_1based):
                data_end_1b = blanks_1based[k] - 1
            else:
                data_end_1b = n_total_rows_1based  # 시트 마지막 행.

            header, data_rows, raw_count = process_subtable(rows, cdid_row_1b, data_end_1b)
            out_name = f"balanceofpayments2025q4_{table_code_lower}_sub{k + 1}.csv"
            out_path = OUT_DIR / out_name
            with out_path.open("w", encoding="utf-8", newline="") as f:
                w = csv.writer(f)
                w.writerow(header)
                w.writerows(data_rows)

            n_data_cols = len(header) - 1  # 시점 컬럼 제외한 CDID 컬럼 수.
            log_rows.append(
                {
                    "sheet": sn,
                    "sub_table": str(k + 1),
                    "out_file": out_name,
                    "n_data_rows": len(data_rows),
                    "n_data_cols": n_data_cols,
                    "raw_cell_count": raw_count,
                    "skipped_reason": "",
                }
            )
            out_files_total += 1
            cells_total += raw_count
    wb.close()

    # 로그 CSV 작성(검증·재현용).
    log_fields = [
        "sheet",
        "sub_table",
        "out_file",
        "n_data_rows",
        "n_data_cols",
        "raw_cell_count",
        "skipped_reason",
    ]
    with LOG_CSV.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=log_fields)
        w.writeheader()
        w.writerows(log_rows)

    print(
        f"OK out_files={out_files_total} cells_total={cells_total} "
        f"log={LOG_CSV.relative_to(ROOT)}"
    )


if __name__ == "__main__":
    main()
