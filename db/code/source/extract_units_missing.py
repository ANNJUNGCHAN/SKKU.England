"""ONS BoP xlsx 시트별 단위/결측 카탈로그 추출.

목적: db/source/balanceofpayments2025q4.xlsx를 read-only로 열고,
  (1) 시트/부표 메타 영역(cdid 행 위 7행)에서 단위 진술 발췌,
  (2) 데이터 영역의 모든 셀을 순회해 결측 표기 등장 빈도/위치 기록.
  12회차/13회차와 (sheet, sub_table) 키로 join 가능.

출력: background/note/15_units.csv, 15_missing.csv
원본 보존: openpyxl read_only=True. 결측을 NA/0 등으로 치환하지 않음.
"""
from __future__ import annotations
import csv
import re
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
SRC = ROOT / "db" / "source" / "balanceofpayments2025q4.xlsx"
NOTE = ROOT / "background" / "note"
INV = NOTE / "12_xlsx_sheet_inventory.csv"
OUT_UNITS = NOTE / "15_units.csv"
OUT_MISS = NOTE / "15_missing.csv"

# CDID 행 인식: 13회차 스크립트와 동일.
PURE = re.compile(r"^[A-Z0-9]{4}$")
PREF = re.compile(r"^\s*-\s*[A-Z0-9]{4}\s*$")
LET = re.compile(r"[A-Z]")
NUM_RE = re.compile(r"^[+-]?[\d,]+(\.\d+)?$")

POUND = chr(163)
# (정규식, 정규화 라벨, scale_factor) 삼중쌍.
# scale_factor 정보는 카탈로그용. 실제 환산은 적재 단계에서 수행.
UNIT_NORMALIZATION = [
    (re.compile(POUND + r"\s*million", re.I), "GBP_million", 1e6),
    (re.compile(r"GBP\s*million", re.I), "GBP_million", 1e6),
    (re.compile(POUND + r"\s*billion", re.I), "GBP_billion", 1e9),
    (re.compile(r"GBP\s*billion", re.I), "GBP_billion", 1e9),
    (re.compile(r"%\s*of\s*GDP", re.I), "pct_of_GDP", None),
    (re.compile(r"per\s*cent\s*of\s*GDP", re.I), "pct_of_GDP", None),
    (re.compile(r"percentage", re.I), "pct", None),
]

UNIT_PAT = re.compile(
    POUND + r"\s*(million|billion|m|bn)|GBP\s*(million|billion)|"
    r"%\s*of\s*GDP|per\s*cent\s*of\s*GDP|seasonally\s*adjusted|"
    r"not\s*seasonally\s*adjusted|net\s*debits|debits\s*=|"
    r"credits\s*\(|net\s*credits|credit|debit",
    re.I,
)

# ONS Service Manual / Government Analysis Function 권장 의미(10회차 §발췌 #9).
MISSING_MEANINGS = {
    "x": "ONS confidential / suppressed (not released; cf Service Manual [c])",
    "..": "ONS legacy: not available (Service Manual now discourages; cf [x])",
    "-": "ONS legacy: nil or true zero (Service Manual now discourages; cf [z])",
    "": "Empty cell - series start before/after data window or non-applicable",
    "[c]": "Confidential (Government Analysis Function symbol)",
    "[x]": "Not available (GAF symbol)",
    "[z]": "Not applicable (GAF symbol)",
    "[low]": "Rounded to zero (GAF symbol)",
    "n/a": "Not applicable (deprecated; ambiguous per ONS Service Manual)",
    "NA": "Not available/applicable (deprecated; ambiguous per ONS Service Manual)",
}

def is_cdid(raw):
    if not isinstance(raw, str):
        return False
    s = raw.strip()
    if not (PURE.match(s) or PREF.match(raw)):
        return False
    code = re.sub(r"[^A-Z0-9]", "", raw)
    if code == "CDID":
        return False
    return bool(LET.search(code))

def normalize_unit(text):
    for pat, label, scale in UNIT_NORMALIZATION:
        if pat.search(text):
            return (label, scale)
    return ("unknown", None)

def detect_missing_marker(value):
    """결측 표기 분류. 숫자(int/float, 또는 숫자 문자열)는 데이터로 간주(None)."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return None
    s = str(value).strip()
    if s == "":
        return ""
    if NUM_RE.match(s):
        return None
    return s

def main():
    inv = {}
    with INV.open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            inv[row["sheet_name"]] = row
    wb = load_workbook(SRC, read_only=True, data_only=True)
    units_rows = []
    missing_counter = defaultdict(lambda: defaultdict(lambda: {"count": 0, "sample": ""}))
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        meta = inv.get(sn, {})
        cdid_row_indices = []
        for i, row in enumerate(rows):
            cnt = sum(1 for v in row if isinstance(v, str) and is_cdid(v))
            if cnt >= 2:
                cdid_row_indices.append(i)
        if not cdid_row_indices:
            sub_ranges = [(1, "0", 0, len(rows) - 1)]
        else:
            sub_ranges = []
            for k, ridx in enumerate(cdid_row_indices):
                start_data = ridx + 1
                end_data = (cdid_row_indices[k + 1] - 1) if k + 1 < len(cdid_row_indices) else len(rows) - 1
                sub_ranges.append((k + 1, str(k + 1), start_data, end_data))
        for k, (sub_num, sub_label, start_data, end_data) in enumerate(sub_ranges):
            if cdid_row_indices:
                ridx = cdid_row_indices[k]
                meta_start = max(0, ridx - 7)
                meta_end = ridx
            else:
                meta_start, meta_end = 0, min(7, len(rows))
            unit_texts = []
            for ri in range(meta_start, meta_end):
                for cv in rows[ri]:
                    if isinstance(cv, str):
                        s = cv.strip()
                        if not s:
                            continue
                        if UNIT_PAT.search(s):
                            unit_texts.append(s[:200])
            if not unit_texts:
                fallback = meta.get("unit_text", "")
                if fallback and fallback != "n/a":
                    unit_texts = [fallback]
            seen = set()
            wrote_any = False
            for ut in unit_texts:
                if ut in seen:
                    continue
                seen.add(ut)
                label, scale = normalize_unit(ut)
                units_rows.append({"sheet": sn, "sub_table": sub_label, "unit_text_raw": ut, "unit_normalized": label, "scale_factor": "" if scale is None else str(int(scale)), "notes": meta.get("notes", "")})
                wrote_any = True
            if not wrote_any:
                units_rows.append({"sheet": sn, "sub_table": sub_label, "unit_text_raw": "(no unit statement found in metadata rows)", "unit_normalized": "unknown", "scale_factor": "", "notes": meta.get("notes", "")})
            for ri in range(start_data, end_data + 1):
                row = rows[ri]
                for ci, val in enumerate(row):
                    if ci == 0:
                        continue
                    marker = detect_missing_marker(val)
                    if marker is None:
                        continue
                    key = (sn, sub_label)
                    bucket = missing_counter[key][marker]
                    bucket["count"] += 1
                    if not bucket["sample"]:
                        col_letter = get_column_letter(ci + 1)
                        bucket["sample"] = col_letter + str(ri + 1)
    wb.close()
    NOTE.mkdir(parents=True, exist_ok=True)
    unit_fields = ["sheet", "sub_table", "unit_text_raw", "unit_normalized", "scale_factor", "notes"]
    with OUT_UNITS.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=unit_fields)
        w.writeheader()
        w.writerows(units_rows)
    miss_fields = ["sheet", "sub_table", "missing_marker", "count", "sample_position", "suggested_meaning"]
    miss_rows = []
    for (sn, sub), markers in sorted(missing_counter.items()):
        for marker, info in sorted(markers.items(), key=lambda kv: -kv[1]["count"]):
            meaning = MISSING_MEANINGS.get(marker, "Non-numeric token (inspect manually)")
            display_marker = marker if marker != "" else "(empty)"
            miss_rows.append({"sheet": sn, "sub_table": sub, "missing_marker": display_marker, "count": info["count"], "sample_position": info["sample"], "suggested_meaning": meaning})
    with OUT_MISS.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=miss_fields)
        w.writeheader()
        w.writerows(miss_rows)
    uniq_units = len({r["unit_normalized"] for r in units_rows})
    uniq_markers = len({r["missing_marker"] for r in miss_rows})
    print("OK units_rows=" + str(len(units_rows)) + " unique_unit_labels=" + str(uniq_units) + " missing_rows=" + str(len(miss_rows)) + " unique_markers=" + str(uniq_markers))

if __name__ == "__main__":
    main()
