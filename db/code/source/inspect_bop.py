"""
inspect_bop.py — 영국 BoP 원본 xlsx 시트 구조 점검 스크립트
================================================================

[목적]
    `db/source/balanceofpayments2025q4.xlsx` 원본 파일의 모든 시트에 대해
    상위 25행을 그대로 출력해, 시트별 헤더 구조·단위·부표 시작 위치를
    한눈에 파악할 수 있게 합니다. REPORT.md 작성·tidy 변환 설계의
    1차 점검 도구로 사용합니다.

[전제]
    - 원본 파일은 `db/source/<항목><yyyy><기간>.xlsx` 패턴을 따릅니다.
    - 본 스크립트는 원본을 **읽기 전용**으로만 접근합니다. 값 변경 금지.
    - 가상환경: 저장소 루트 `env/` (Windows: `env\\Scripts\\python.exe`).
      필수 패키지: `openpyxl` (requirements.txt 동기화 완료).

[사용법]
    1) 가상환경 활성화 (PowerShell)
         > .\\env\\Scripts\\Activate.ps1
       또는 Git Bash
         $ source env/Scripts/activate

    2) 저장소 루트에서 실행
         (env) > python db/code/source/inspect_bop.py

    3) 출력 리다이렉션 (대용량이므로 파일로 저장 권장)
         (env) > python db/code/source/inspect_bop.py > .inspect.txt

[출력 형식]
    file: <원본 절대경로>
    sheets (N): [...]
    ================================================================
    ## sheet: <시트명>
    dimensions: A1:S30  max_row=30  max_col=19
      r  1: ['<셀 A>', '<셀 B>', ...]   ← 각 셀 최대 60자 미리보기
      ...

[입력 변경]
    분석 대상 파일이 바뀌면 아래 SRC 상수를 수정합니다.
    파일명 규칙: `db/source/<항목><yyyy><기간>.xlsx`.

[주의사항]
    - Windows cp949 환경에서 £ 기호 등 비ASCII 출력이 깨질 수 있어
      stdout을 UTF-8로 강제 래핑합니다.
    - openpyxl의 read_only 모드는 `dimensions` 속성을 제공하지 않으므로
      `read_only=False`로 로드합니다(메모리는 더 쓰지만 점검에 충분).
    - 본 스크립트는 **출력 전용**으로, 어떤 파일도 생성·수정하지 않습니다.

[관련 파일]
    - `db/code/source/extract_latest.py` : 분기 시계열 말단 행 추출
    - `db/REPORT.md`                     : 본 스크립트 출력에 기반한 분석
    - `db/CLAUDE.md`                     : `db/source/` 가공 규칙
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl

# Windows 한국어 로캘에서도 £ 등 비ASCII 문자가 깨지지 않도록 stdout을 UTF-8로 강제.
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# 저장소 루트 = 본 파일 기준 3단계 상위 (db/code/source/inspect_bop.py → repo root).
REPO_ROOT = Path(__file__).resolve().parents[3]
SRC = REPO_ROOT / "db" / "source" / "balanceofpayments2025q4.xlsx"


def main() -> None:
    # data_only=True 로 수식 결과값을 읽어옵니다(원본 셀 자체는 변경하지 않음).
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=False)

    print(f"file: {SRC}")
    print(f"sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
    print("=" * 80)

    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n## sheet: {name}")
        print(
            f"dimensions: {ws.dimensions}  "
            f"max_row={ws.max_row}  max_col={ws.max_column}"
        )

        # 시트별로 상위 25행만 미리보기. 부표 헤더(보통 r6–r8)와 본문 시작 위치 확인 목적.
        rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), values_only=True))
        for i, row in enumerate(rows, 1):
            trimmed = [c for c in row if c is not None]
            if not trimmed:
                print(f"  r{i:>3}: <blank>")
                continue
            preview = [str(c)[:60] for c in row[:12]]
            print(f"  r{i:>3}: {preview}")


if __name__ == "__main__":
    main()
