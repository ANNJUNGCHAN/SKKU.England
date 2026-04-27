"""ONS BoP 분기 통계 불러틴 표(xlsx)에서 CDID 사전을 추출하는 절차.

목적
    `db/source/balanceofpayments2025q4.xlsx`의 본표 시트들을 읽어, 각 시트·부표·
    컬럼 위치별로 CDID(ONS Time Series 4자 식별자)를 모아 `background/note/
    13_cdid_dictionary.csv`로 저장한다. 12회차 시트 인벤토리(`12_xlsx_sheet_
    inventory.csv`)에서 CDID 행이 8행으로 일관됨이 확인된 본표만 대상으로 한다.

입력
    - `db/source/balanceofpayments2025q4.xlsx` (읽기 전용 접근)
    - `background/note/12_xlsx_sheet_inventory.csv` (시트 메타·표 코드·분류 출처)

출력
    - `background/note/13_cdid_dictionary.csv` (10 컬럼: sheet, sub_table,
      column_position, column_label, cdid, sign_prefix, unit, table_code,
      classification, notes)
    - 표준출력에 추출 행 수·고유 CDID 수·sign_prefix 갯수 요약 한 줄.

부작용
    - `background/note/` 디렉터리가 없으면 생성한다.
    - 원본 xlsx는 절대 수정하지 않는다(openpyxl을 read_only=True로 호출).

실행 방법
    env/Scripts/python.exe db/code/source/extract_cdid.py
"""
from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook

# 저장소 루트 기준 경로 고정. 실행 위치와 무관하게 동일 산출.
ROOT = Path(__file__).resolve().parents[3]
SRC = ROOT / "db" / "source" / "balanceofpayments2025q4.xlsx"
NOTE = ROOT / "background" / "note"
INV = NOTE / "12_xlsx_sheet_inventory.csv"
OUT_CSV = NOTE / "13_cdid_dictionary.csv"

# CDID 후보 정규식: 영숫자 4자(예: BOKI, L87S, MU7L). Table_C 등 숫자 포함 코드도 포착.
PURE = re.compile(r"^[A-Z0-9]{4}$")
# 부호 반전 규칙 표기: 앞에 공백+마이너스가 붙은 CDID(예: " -HJYM"). Notes의 Table A
# note 1에 따라 적재 시 값 부호를 뒤집어야 함.
PREF = re.compile(r"^\s*-\s*[A-Z0-9]{4}\s*$")
# 영문이 한 글자라도 있어야 CDID로 인정. 숫자만 4자(예: 2025)는 연도이므로 제외.
LET = re.compile(r"[A-Z]")
# 라벨 셀 "CDID" 자체는 추출에서 제외.
EXCL = {"CDID"}


def cs(v):
    """엑셀 셀 값을 문자열로 정규화(None은 빈 문자열)."""
    return "" if v is None else str(v)


def is_cdid(raw):
    """원시 셀 문자열이 CDID인지 판정해 4자 코드(부호 prefix 제거)를 반환.

    CDID가 아니면 None.
    """
    s = raw.strip()
    if not (PURE.match(s) or PREF.match(raw)):
        return None
    # 부호 prefix·공백을 제거한 4자 코드만 남김.
    code = re.sub(r"[^A-Z0-9]", "", raw)
    if code in EXCL:
        return None
    if not LET.search(code):
        return None
    return code


def main():
    # 12회차 인벤토리에서 CDID 행 = 8행으로 식별된 본표 시트만 대상.
    inv = {}
    with INV.open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row["cdid_row"] == "8":
                inv[row["sheet_name"]] = row

    # 원본은 절대 수정하지 않도록 read_only=True. data_only=True로 수식이 아닌 값을 읽음.
    wb = load_workbook(SRC, read_only=True, data_only=True)
    out_rows = []
    for sn in wb.sheetnames:
        if sn not in inv:
            continue
        ws = wb[sn]
        # 시트 전체를 메모리로 한 번에 끌어올린 뒤 인덱싱(부표 경계 탐색에 필요).
        rows = [[cs(c) for c in r] for r in ws.iter_rows(values_only=True)]

        # CDID 행을 모두 찾는다(부표가 여러 개면 시트 안에 여러 CDID 행이 등장).
        # 한 행에 CDID가 2개 이상이어야 진짜 CDID 행으로 인정(라벨 충돌 회피).
        cdid_row_indices = []
        for i, row in enumerate(rows):
            if sum(1 for v in row if is_cdid(v) is not None) >= 2:
                cdid_row_indices.append(i)

        meta = inv[sn]
        for k, ridx in enumerate(cdid_row_indices):
            cdid_row = rows[ridx]
            # 컬럼 라벨 행을 찾는다: CDID 행 바로 위 1~3행 중 비어 있지 않은 첫 행.
            header = []
            for back in (1, 2, 3):
                if ridx - back < 0:
                    continue
                cand = rows[ridx - back]
                if any(v.strip() for v in cand):
                    header = cand
                    break

            sub = k + 1  # 부표 식별자(시트 안에서 1부터 순서대로).
            for cp, raw in enumerate(cdid_row):
                cd = is_cdid(raw)
                if cd is None:
                    continue
                pref = bool(PREF.match(raw))
                lab = header[cp] if cp < len(header) else ""
                # 줄바꿈을 공백으로 평탄화해 CSV 단일 셀 안에 안전하게 저장.
                lab = lab.strip().replace(chr(10), " ")
                out_rows.append({
                    "sheet": sn,
                    "sub_table": str(sub),
                    "column_position": str(cp + 1),
                    "column_label": lab,
                    "cdid": cd,
                    "sign_prefix": "true" if pref else "false",
                    "unit": meta.get("unit_text", ""),
                    "table_code": meta.get("table_code", ""),
                    "classification": meta.get("classification", ""),
                    "notes": meta.get("notes", ""),
                })
    wb.close()

    fields = [
        "sheet", "sub_table", "column_position", "column_label",
        "cdid", "sign_prefix", "unit", "table_code", "classification", "notes",
    ]
    NOTE.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(out_rows)

    uniq = len({r["cdid"] for r in out_rows})
    sign = sum(1 for r in out_rows if r["sign_prefix"] == "true")
    print(f"OK rows={len(out_rows)} unique={uniq} sign={sign}")


if __name__ == "__main__":
    main()
