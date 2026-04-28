"""Phase 2.4 검증 스크립트 (§2.4 첫 번째 항목).

목적
----
db/CHECKLIST.md §2.4 "원본 셀 수와 가공 후 CSV 셀 수의 합이 일치"를 자동 점검한다.

방법
----
1. 원본 측 셀 수: ONS xlsx의 본표 17 시트에서 부표별 데이터 영역(CDID 행 +
   데이터 행) × CDID 컬럼 수 + 시점 컬럼 = (n_data_rows + 1) × (n_cdid_cols + 1).
   메타 시트 3개는 별도 CSV 산출하지 않으므로 검증 대상 외.
2. 가공 측 셀 수: 본 ETL이 생성한 63개 부표 CSV의 (행 × 열) 합계.
3. 두 수치를 비교해 일치 여부를 출력. 추가로 결측 마커 보존(Table_C `x` 360건)도
   재확인.

원본 보존
----------
xlsx는 read_only로만 접근. 결측 또는 부호 prefix 등은 어떤 단계에서도 변경하지
않는다.

재현
----
env/Scripts/python.exe db/code/source/verify_phase2_1.py
"""
from __future__ import annotations

import csv
import re
from pathlib import Path
from openpyxl import load_workbook

# 저장소 루트.
ROOT = Path(__file__).resolve().parents[3]
SRC = ROOT / "db" / "source" / "balanceofpayments2025q4.xlsx"
INV = ROOT / "db" / "data" / "_inventory" / "15_master_inventory.csv"
LOG = ROOT / "db" / "data" / "_etl_log" / "phase2_1_split_log.csv"
DATA_DIR = ROOT / "db" / "data"

# 13회차·12회차 동일 CDID 인식 정규식.
PURE = re.compile(r"^[A-Z0-9]{4}$")
PREF = re.compile(r"^\s*-\s*[A-Z0-9]{4}\s*$")
LET = re.compile(r"[A-Z]")


def is_cdid(raw: object) -> bool:
    if not isinstance(raw, str):
        return False
    s = raw.strip()
    if not (PURE.match(s) or PREF.match(raw)):
        return False
    code = re.sub(r"[^A-Z0-9]", "", raw)
    if code == "CDID":
        return False
    return bool(LET.search(code))


def count_x_in_csv(path: Path) -> int:
    """CSV에서 `x` 셀(소문자 단일 문자) 개수를 카운트."""
    n = 0
    with path.open(encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader)  # 헤더 제외.
        for row in reader:
            for v in row:
                if v == "x":
                    n += 1
    return n


def main() -> None:
    """검증 진입점."""
    # 원본 측 카운트.
    with INV.open(encoding="utf-8") as f:
        masters = list(csv.DictReader(f))
    wb = load_workbook(SRC, read_only=True, data_only=True)

    raw_cells_total = 0
    raw_x_total = 0
    raw_cells_per_sheet: dict[str, int] = {}
    for m in masters:
        if m["classification_code"] == "meta_notes":
            continue
        sn = m["sheet"]
        ws = wb[sn]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        cdid_rows = [int(x) - 1 for x in m["all_cdid_rows"].split(";") if x]
        blanks = [int(x) - 1 for x in m["blank_row_positions"].split(";") if x]
        raw_cells_sheet = 0
        for k, ridx in enumerate(cdid_rows):
            sub_end_idx = (blanks[k] - 1) if k < len(blanks) else len(rows) - 1
            cdid_row = rows[ridx]
            cdid_cols = [ci for ci, v in enumerate(cdid_row) if isinstance(v, str) and is_cdid(v)]
            n_cols = len(cdid_cols)
            # 부표 셀 수 = (CDID 헤더 행 1 + 데이터 행) × (시점 컬럼 1 + CDID 컬럼).
            n_data_rows = sub_end_idx - ridx  # CDID 행(+1) 다음 데이터 행 ridx+1 ~ sub_end_idx
            n_total_rows = 1 + n_data_rows  # CDID 헤더 + 데이터.
            raw_cells_sheet += n_total_rows * (1 + n_cols)
            # 결측 `x` 카운트.
            for ri in range(ridx + 1, sub_end_idx + 1):
                if ri >= len(rows):
                    break
                row = rows[ri]
                for ci in cdid_cols:
                    if ci < len(row):
                        v = row[ci]
                        if isinstance(v, str) and v.strip() == "x":
                            raw_x_total += 1
        raw_cells_per_sheet[sn] = raw_cells_sheet
        raw_cells_total += raw_cells_sheet
    wb.close()

    # 가공 측 카운트.
    with LOG.open(encoding="utf-8") as f:
        log_rows = list(csv.DictReader(f))
    csv_cells_total = 0
    csv_x_total = 0
    csv_files = 0
    csv_cells_per_sheet: dict[str, int] = {}
    for r in log_rows:
        if not r["out_file"]:
            continue
        path = DATA_DIR / r["out_file"]
        with path.open(encoding="utf-8") as f:
            reader = csv.reader(f)
            n_rows = 0
            n_cols = 0
            for i, row in enumerate(reader):
                if i == 0:
                    n_cols = len(row)
                n_rows += 1
        cells = n_rows * n_cols
        csv_cells_total += cells
        csv_files += 1
        sn = r["sheet"]
        csv_cells_per_sheet[sn] = csv_cells_per_sheet.get(sn, 0) + cells
        # 결측 `x` 보존 카운트.
        csv_x_total += count_x_in_csv(path)

    # 결과 출력.
    print("Phase 2.4 검증 결과")
    print("=" * 60)
    print(f"원본 본표 셀 수 합계: {raw_cells_total:,}")
    print(f"가공 CSV 셀 수 합계:  {csv_cells_total:,}")
    print(f"차이: {csv_cells_total - raw_cells_total:+,}")
    print(f"가공 CSV 파일 수:     {csv_files}")
    print()
    print(f"원본 `x` 결측 마커 수: {raw_x_total}")
    print(f"가공 `x` 결측 마커 수: {csv_x_total}")
    print(f"`x` 보존 동일성: {'OK' if raw_x_total == csv_x_total else 'FAIL'}")
    print()
    print("시트별 셀 수 비교 (불일치 시트만 표시):")
    mismatches = 0
    for sn in raw_cells_per_sheet:
        rc = raw_cells_per_sheet[sn]
        cc = csv_cells_per_sheet.get(sn, 0)
        if rc != cc:
            print(f"  {sn}: raw={rc:,} csv={cc:,} diff={cc-rc:+,}")
            mismatches += 1
    if mismatches == 0:
        print("  (모든 시트 일치)")
    print()
    overall_ok = (raw_cells_total == csv_cells_total) and (raw_x_total == csv_x_total) and (mismatches == 0)
    print(f"Phase 2.4 §1 자동 점검: {'PASS' if overall_ok else 'FAIL'}")


if __name__ == "__main__":
    main()
